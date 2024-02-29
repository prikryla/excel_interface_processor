import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

def process_files():
    try:
        # Get the paths of file1 and file2
        file1_path = file1_entry.get()
        file2_path = file2_entry.get()

        # Load both Excel files
        file1 = load_workbook(file1_path)
        file2 = load_workbook(file2_path)

        # Assuming the cell is in the first sheet for both files
        sheet1 = file1.active
        sheet2 = file2.active

        # Find the last row with data in file1 column C
        last_row_file1 = sheet1.max_row
        for row_num in range(4, last_row_file1 + 1):
            cell_value = sheet1[f"C{row_num}"].value
            if not cell_value:
                last_row_file1 = row_num - 1
                break

        # Define the range of rows to process in file2
        start_row_file2 = 4
        end_row_file2 = last_row_file1

        # Get the range of cells in file1 (from C4 to the last row with data)
        cells_file1 = sheet1[f"C{start_row_file2}:C{end_row_file2}"]

        # Iterate over cells in file1
        for cell_tuple in cells_file1:
            for cell_file1 in cell_tuple:
                value_to_match = str(cell_file1.value)
                fill_color_to_copy = cell_file1.fill.start_color
                
                # Iterate over cells in file2 for each cell in file1
                for row_num_file2 in range(start_row_file2, end_row_file2 + 1):
                    cell_ref_file2 = f"C{row_num_file2}"  # Cell reference like C4, C5, ..., C22
                    
                    # Get the cell in file2
                    cell_file2 = sheet2[cell_ref_file2]
                    
                    # Compare the value with the value from file1
                    if str(cell_file2.value) == value_to_match:
                        # Create a new PatternFill object with the fill color from file1
                        fill_file2 = PatternFill(start_color=fill_color_to_copy, end_color=fill_color_to_copy, fill_type="solid")

                        # Set the fill color of the corresponding cell in file2
                        cell_file2.fill = fill_file2
                        
                        # Copy the value from cell M in file1 to the corresponding cell in file2
                        cell_M_file1 = sheet1[f"M{cell_file1.row}"]
                        cell_M_file2 = sheet2[f"M{cell_ref_file2[1:]}"]
                        cell_M_file2.value = cell_M_file1.value
                
                # No match found for the current value in file1, continue to the next value

        # Save the updated file2
        output_file_path = os.path.join(os.path.dirname(file2_path), "updated_file2.xlsx")
        file2.save(output_file_path)
        
        success_label.config(text="Files processed successfully!")
        print("Updated file saved to:", output_file_path)      
          
    except Exception as e:
        print(e)
        error_label.config(text="Something went wrong! Check out the old file and try again!")


def browse_file(entry_widget):
    filename = filedialog.askopenfilename()
    entry_widget.delete(0, tk.END)
    entry_widget.insert(0, filename)

# Create GUI
root = tk.Tk()
root.title("Interface processing helper")

# Text explaining the purpose of the tool
info_text = """This is a small tool that helps you with processing interface Excel files. 
Select a day old Excel file first and then the actual file. 
The actual file will be processed and updated based on data from the previous old file. NOTE: This is just the first version of this tool. Any feedback or error report is appreciated!"""
info_label = tk.Label(root, text=info_text, wraplength=400)
info_label.grid(row=0, columnspan=3, padx=5, pady=5)

# Input field and label for file1 path
file1_label = tk.Label(root, text="Select File 1:")
file1_label.grid(row=1, column=0, padx=5, pady=5)
file1_entry = tk.Entry(root, width=50)
file1_entry.grid(row=1, column=1, padx=5, pady=5)
file1_button = tk.Button(root, text="Browse", command=lambda: browse_file(file1_entry))
file1_button.grid(row=1, column=2, padx=5, pady=5)

# Input fields for file2 path
file2_label = tk.Label(root, text="Select File 2:")
file2_label.grid(row=2, column=0, padx=5, pady=5)
file2_entry = tk.Entry(root, width=50)
file2_entry.grid(row=2, column=1, padx=5, pady=5)
file2_button = tk.Button(root, text="Browse", command=lambda: browse_file(file2_entry))
file2_button.grid(row=2, column=2, padx=5, pady=5)

# Process button
process_button = tk.Button(root, text="Process Files", command=process_files)
process_button.grid(row=4, column=1, padx=5, pady=5)

# Label to display success message
success_label = tk.Label(root, text="", fg="green")
success_label.grid(row=5, columnspan=3, padx=5, pady=5)

# Label to display error message
error_label = tk.Label(root, text="", fg="red")
error_label.grid(row=6, columnspan=3, padx=5, pady=5)

root.mainloop()
